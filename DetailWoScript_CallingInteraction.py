Filepath = "C:\\Users\\v-hcyclewala\\OneDrive - Microsoft\\V Desktop\\PythonAutomation\\"
fileName = "Links"
Inputpath = Filepath+fileName+".xlsx"
print(Inputpath)

import pandas as pd
# df = pd.read_excel(Inputpath)
# Column = df.columns.get_loc('Link')
# Column = chr(Column+65)
# print(Column)

import openpyxl

wb = openpyxl.load_workbook(Inputpath)
ws = wb.active

Max_Row = ws.max_row
print(Max_Row)
Links = []
for i in range(1, Max_Row+1):
    Index_Segment_Name = 'A'+str(i)
    Link = ws[Index_Segment_Name].value
    Links.append(Link)

import InteractionDetails as Fi
Idf = Fi.FullInteractionDetails(Links)

# import SegmentDetails as FS
# df = FS.FullSegmentDetails(Idf['Segment_Link'])

df1 = pd.DataFrame({"InteractionName": Idf['InteractionName'], "InteractionLinks": Idf['InteractionLinks'], "InteractionDescription": Idf['InteractionDescription'], "Segment_Id": Idf['Segment_Id'], "Segment_Link": Idf['Segment_Link'], "Segment_Name": Idf['Segment_Name'], "State": Idf['State'], "ApprovalState": Idf['ApprovalState'], "PersonaType": Idf['PersonaType'], "InteractionState": Idf['InteractionState'], "closed": Idf['closed'], "LastUpdated": Idf['LastUpdated'], "RampPercentage": Idf['RampPercentage'], "Audience_Experiment_on": Idf['Audience_Experiment_on'], "Outcome": Idf['Outcome'], "Control": Idf['Control'], "TargetPercent": Idf['TargetPercent']})
df2 = pd.DataFrame({"InteractionName": Idf['InteractionName'], "TreatmentName": Idf['TreatmentName'], "ActionName": Idf['ActionName'], "PayloadId": Idf['PayloadId'], 	"PlatformType":Idf['PlatformType'], "Comm_Pref_Topic": Idf['Comm_Pref_Topic'], "Privacy_Comm_Type": Idf['Privacy_Comm_Type'], "Instance_Name":Idf['Instance_Name'], "DataExtension": Idf['DataExtension'], "TimeBtwView": Idf['TimeBtwView']})
df3 = pd.DataFrame({"InteractionName": Idf['InteractionName'], "StartDate": Idf['StartDate'], "EndDate": Idf['EndDate'], "ScheduleType": Idf['ScheduleType'], "ScheduleDay": Idf['ScheduleDay']})
# df4 = pd.DataFrame({"Segment_Id": df['Segment_Id'], "Segment_Link": df['Segment_Link'], "Segment_Name": df['Segment_Name'], "Segment_Description": df['Segment_Description'], "Segment_LineOfBusiness": df['Segment_LineOfBusiness'],"Segment_Tags": df['Segment_Tags'], "Segment_Product": df['Segment_Product'], "Segment_UserType": df['Segment_UserType'], "Segment_SourceType": df['Segment_SourceType'], "Segment_QueryType": df['Segment_QueryType'], "Segment_DelayPublish": df['Segment_DelayPublish']})
# df5 = pd.DataFrame({"Segment_Name": df['Segment_Name'], "Date1": df['Date1'], "Job_Status1": df['Job_Status1'], "Counts1": df['Counts1']})

SheetDetails = {"Interaction_Detail": df1, "Treatment_Detail": df2, "Schedule_Detail": df3}
# , "SegmentDetails": df4, "InSegmentDetails": df5}

# Segment_Id Segment_Description
# writer = pd.ExcelWriter(r"C:\Users\v-hcyclewala\OneDrive - Microsoft\V Desktop\PythonHandson\Forza.xlsx", engine="xlsxwriter")

from datetime import datetime

current_datetime = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
str_current_datetime = str(current_datetime)
OutputPath = Filepath+"Output"+str_current_datetime+".xlsx"

with pd.ExcelWriter(OutputPath) as writer:
    for sheet_Name in SheetDetails.keys():
        SheetDetails[sheet_Name].to_excel(writer, sheet_name=sheet_Name, index=False)