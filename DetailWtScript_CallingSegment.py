Filepath = "C:\\Users\\v-hcyclewala\\OneDrive - Microsoft\\V Desktop\\PythonAutomation\\"
fileName = "Links"
Inputpath = Filepath+fileName+".xlsx"
print(Inputpath)

import pandas as pd
# df = pd.read_excel(Inputpath)
# Column = df.columns.get_loc('Link')
# Column = chr(Column+65)
# print(Column)

import openpyxl

wb = openpyxl.load_workbook(Inputpath)
ws = wb.active

Max_Row = ws.max_row
print(Max_Row)
Links = []
for i in range(1, Max_Row+1):
    Index_Segment_Name = 'A'+str(i)
    Link = ws[Index_Segment_Name].value
    Links.append(Link)

import SegmentDetails as FS
df = FS.FullSegmentDetails(Links, Filepath)

df4 = pd.DataFrame({"Segment_Id": df['Segment_Id'], "Segment_Link": df['Segment_Link'], "Segment_Name": df['Segment_Name'], "Segment_Description": df['Segment_Description'], "Segment_LineOfBusiness": df['Segment_LineOfBusiness'],"Segment_Tags": df['Segment_Tags'], "Segment_Product": df['Segment_Product'], "Segment_UserType": df['Segment_UserType'], "Segment_SourceType": df['Segment_SourceType'], "Segment_QueryType": df['Segment_QueryType'], "Segment_DelayPublish": df['Segment_DelayPublish']})
df5 = pd.DataFrame({"Segment_Name": df['Segment_Name'], "Script_WhichVC": df['Script_WhichVC'], "Script_GetCounts": df['Script_GetCounts'],  "Tag_AudienceId": df['Tag_AudienceId'], "Tag_Campaign": df['Tag_Campaign'], "Tag_Freq": df['Tag_Freq'], "Script_CollectionId": df['Script_CollectionId'], "Script_ProgramId": df['Script_ProgramId'], "Script_AudienceId": df['Script_AudienceId'], "Date1": df['Date1'], "Job_Status1": df['Job_Status1'], "Counts1": df['Counts1']})
# , "Date2": df['Date2'], "Job_Status2": df['Job_Status2'], "Counts2": df['Counts2'], "Date3": df['Date3'], "Job_Status3": df['Job_Status3'], "Counts3": df['Counts3']
SheetDetails = {"SegmentDetails": df4, "InSegmentDetails": df5}
# Segment_Id Segment_Description
# writer = pd.ExcelWriter(r"C:\Users\v-hcyclewala\OneDrive - Microsoft\V Desktop\PythonHandson\Forza.xlsx", engine="xlsxwriter")

from datetime import datetime

current_datetime = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
str_current_datetime = str(current_datetime)
OutputPath = Filepath+"SegmentDetailsWtScript"+str_current_datetime+".xlsx"

with pd.ExcelWriter(OutputPath) as writer:
    for sheet_Name in SheetDetails.keys():
        SheetDetails[sheet_Name].to_excel(writer, sheet_name=sheet_Name, index=False)