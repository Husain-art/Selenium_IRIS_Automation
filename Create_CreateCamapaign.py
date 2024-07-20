from selenium import webdriver
from selenium.webdriver import Keys
from selenium .webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.edge.service import Service

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from selenium.common.exceptions import NoSuchElementException

from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
import time
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from selenium.webdriver.remote import webelement
from selenium.webdriver.support.wait import WebDriverWait


# //div[@data-test= "name"]//input
IRISError_Xpath = '//div[@data-test="Error report"]'

def create_Campaign(x1, x2, x3, x4, x5, x6, x7, x8, x9, x10, x11, x12, x13, x14, x15, x16, x17, x18, x19, x20, x21):
    website = 'https://studio.iris.microsoft.com/#/campaigns/create'

    service = Service(executable_path=r"C:\Users\v-hcyclewala\Downloads\edgedriver_win64\msedgedriver.exe")
    driver = webdriver.Edge(service=service)
    driver.get(website)
    driver.maximize_window()

    loginButton = driver.find_element(by='xpath', value='//button[@class = "ms-Button ms-Button--primary root-83"]')
    loginButton.click()
    time.sleep(3)

    CampaignName = WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, '//div[@data-test= "name"]//input')))
    CampaignName.send_keys(x1)

    lineOfBusiness = driver.find_element(by='xpath', value='//div[@data-test= "lineOfBusiness"]/div/div/div')
    lineOfBusiness.click()

    lineOfBusinessValue_XpathText = '//button[@title=\"'+x2+'\"]'
    # print(ProductValue_XpathText)
    lineOfBusinessButton = driver.find_element(by='xpath', value=lineOfBusinessValue_XpathText)
    lineOfBusinessButton.click()

    Description = driver.find_element(by='xpath', value='//div[@data-test= "description"]//textarea')
    Description.send_keys(x3)

    CampaignOwningProduct = driver.find_element(by='xpath', value='//div[@data-test= "product"]/div/div')
    CampaignOwningProduct.click()

    CampaignOwningProductValue_XpathText = '//button[@title=\"'+x4+'\"]'
    # print(ProductValue_XpathText)
    CampaignOwningProductButton = driver.find_element(by='xpath', value=CampaignOwningProductValue_XpathText)
    CampaignOwningProductButton.click()

    lifecycle = driver.find_element(by='xpath', value='//div[@data-test= "lifecycle"]/div/div')
    lifecycle.click()

    lifecycleValue_XpathText = '//button[@title=\"' +x5+ '\"]'
    # print(ProductValue_XpathText)
    lifecycleButton = driver.find_element(by='xpath', value=lifecycleValue_XpathText)
    lifecycleButton.click()

    InteractionName = WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, '//div[@data-test="interactions"]//div[@data-test= "name"]//input')))
    InteractionName.send_keys(x6)

    # //div[@data-test="interactions"]//div[@data-test= "description"]//textarea
    InteractionDescription = driver.find_element(by='xpath', value='//div[@data-test="interactions"]//div[@data-test= "description"]//textarea')
    InteractionDescription.send_keys(x7)

    # //h1[text() = 'Select outcomes']
    try:
        WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, "//h1[text()='Select outcomes']")))
        action = ActionChains(driver)
        Outcome = WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, '//div[@data-test="outcomes"]//div[@data-list-index="19"]')))
        action.scroll_to_element(Outcome).perform()

        FocusedProduct = driver.find_element(by='xpath', value='//div[@data-test="productList"]')
        SelectedItemsResult = driver.find_element(by='xpath', value='//div[@data-test="outcomes"][2]')

        action.scroll_to_element(FocusedProduct).perform()
        action.scroll_to_element(SelectedItemsResult).perform()
        action.scroll_to_element(FocusedProduct).perform()
        action.scroll_to_element(SelectedItemsResult).perform()

        Outcome = WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, '//div[@data-test="outcomes"]//div[@data-list-index="39"]')))
        action.scroll_to_element(Outcome).perform()
        Outcome.click()
        action.scroll_to_element(FocusedProduct).perform()
        action.scroll_to_element(SelectedItemsResult).perform()
        action.scroll_to_element(FocusedProduct).perform()
        action.scroll_to_element(SelectedItemsResult).perform()

        Outcome = WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, '//div[@data-test="outcomes"]//div[@data-list-index="59"]')))
        action.scroll_to_element(Outcome).perform()
        Outcome.click()

        action.scroll_to_element(FocusedProduct).perform()
        action.scroll_to_element(SelectedItemsResult).perform()
        action.scroll_to_element(FocusedProduct).perform()
        action.scroll_to_element(SelectedItemsResult).perform()

        Outcome = WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, '//div[@data-test="outcomes"]//div[@data-list-index="54"]')))
        action.scroll_to_element(Outcome).perform()
        Outcome.click()
        # Select Outcome End
        time.sleep(1)
    except NoSuchElementException:
        print("Select Outcome Section is not present")
    # SelectOucome start


    NextButton = driver.find_element(by='xpath', value='//main/div/div/div[2]/div/div/form/div[2]/button[2]')
    NextButton.click()

    UserType = WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, '//div[@data-test="identityType"]/div/div')))
    UserType.click()

    UserTypeValue_XpathText = '//button[@title=\"' + x8 + '\"]'
    # print(ProductValue_XpathText)
    UserTypeButton = driver.find_element(by='xpath', value=UserTypeValue_XpathText)
    UserTypeButton.click()

    SegmentnameButton = WebDriverWait(driver, 360).until(EC.presence_of_element_located((By.XPATH, '//div[@data-item-key="name"]')))
    SegmentnameButton.click()

    SegmentNameInput = driver.find_element(by='xpath', value='//input[@placeholder="Filter by Segment Name"]')
    SegmentNameInput.send_keys(x9)

    Filter = driver.find_element(by='xpath', value='//div/div/div/div/div/div/div/ul/li[7]//button[2]')
    Filter.click()
    time.sleep(2)

    IncludeButton = driver.find_element(by='xpath', value='//div[@data-automation-key="include"]//button')
    IncludeButton.click()
    time.sleep(1)

    NextButton2 = driver.find_element(by='xpath', value='//main/div/div/div[2]/div/div/form/div[2]/button[3]')
    NextButton2.click()

    Control = WebDriverWait(driver, 360).until(EC.presence_of_element_located((By.XPATH, '//div[@data-test="controlGroupSize"]//input')))
    Control.send_keys(x10)

    TreatmentName = driver.find_element(by='xpath', value='//div[@data-automation-key="name"]//input')
    TreatmentName.send_keys(Keys.CONTROL + "A")
    time.sleep(1)
    TreatmentName.send_keys(x11)

    time.sleep(1)

    NextButton2.click()

    ActionButton = WebDriverWait(driver, 360).until(EC.presence_of_element_located((By.XPATH, '//button[@data-test="AddActionIcon"]')))
    ActionButton.click()

    Action_Xpath = '//button[@name=\"' + x12 + '\"]'
    # print(ProductValue_XpathText)
    ActionTake = driver.find_element(by='xpath', value=Action_Xpath)
    ActionTake.click()
    # //div[@data-test="platformType"]//div[@tabindex="0"]
    PlatformType = WebDriverWait(driver, 360).until(EC.presence_of_element_located((By.XPATH, '//div[@data-test="platformType"]')))
    PlatformType.click()

    # //button[@title="SFMC Data Extension"]

    time.sleep(.5)
    PlatformTypeValue_Xpath = '//button[contains(@title,\"' + str(x13) + '\")]'
    # print(ProductValue_XpathText)
    PlatformTypeValue = driver.find_element(by='xpath', value=PlatformTypeValue_Xpath)
    PlatformTypeValue.click()

    # //button[@name="Email"]
    # //div[@data-test="platformType"]//span
    # //button[@title="SFMC Data Extension"]
    # //button[@title="Iris SMTP"]
    # title="Azure Email Orchestrator"
    # title="Adobe Campaign"

    InstanceName = WebDriverWait(driver, 360).until(EC.presence_of_element_located((By.XPATH, '//div[@data-test="instanceName"]//span')))
    InstanceName.click()

    InstanceNameValue_Xpath = '//button[@title=\"' + x14 + '\"]'
    # print(ProductValue_XpathText)
    InstanceNameValue = driver.find_element(by='xpath', value=InstanceNameValue_Xpath)
    InstanceNameValue.click()

    # data-test="instanceName"
    # //button[@title="SFMC_DEVICES_STUDIOS_XBOX_US_PROD"]

    TopicId = WebDriverWait(driver, 360).until(EC.presence_of_element_located((By.XPATH, '//div[@data-test="topicId"]//span')))
    TopicId.click()
    # Xbox Promotional Communications
    TopicIdValue_Xpath = '//button[@title=\"' + x15 + '\"]'
    # print(ProductValue_XpathText)
    TopicIdValue = driver.find_element(by='xpath', value=TopicIdValue_Xpath)
    TopicIdValue.click()

    Interval = driver.find_element(by='xpath', value='//div[@data-test="interval"]//input')
    Interval.send_keys(Keys.CONTROL + "A")
    time.sleep(5)
    Interval.send_keys(x16)
    # //div[@data-test="interval"]//input

    NextButton2.click()

    StartDate = WebDriverWait(driver, 360).until(EC.presence_of_element_located((By.XPATH, '//div[@data-test="startDate"]//input[@placeholder="Choose date"]')))
    StartDate.click()
    time.sleep(.5)
    year = x17.strftime("%Y")
    print(year)
    IRISYear = driver.find_element(by='xpath', value='//div[contains(@class, "ms-DatePicker-monthPicker monthPicker")]//div[contains(@class, "ms-DatePicker-header header")]/div').text
    print("Before Increament", IRISYear)

    while IRISYear != year:
        ClickNextYear = driver.find_element(by='xpath', value='//div[contains(@class, "ms-DatePicker-yearComponents")]//button[contains(@aria-label, "Go to next year")]')
        ClickNextYear.click()
        time.sleep(.5)
        IRISYear = driver.find_element(by='xpath', value='//div[contains(@class, "ms-DatePicker-monthPicker monthPicker")]//div[contains(@class, "ms-DatePicker-header header")]/div').text
        print(IRISYear)
    print("AfterIncreament", IRISYear)

    month = x17.month
    import calendar
    MonthName = calendar.month_name[month]
    print("MonthName", MonthName)

    Month_Xpath = '//div[contains(@class, "ms-DatePicker-monthPicker monthPicker")]//div[3]//button[contains(@aria-label, \"' + MonthName + '\")]'
    # print(ProductValue_XpathText)
    ClickMonth = driver.find_element(by='xpath', value=Month_Xpath)
    ClickMonth.click()
    time.sleep(.5)

    Day = x17.day
    print("Day", Day)

    Month_Day = MonthName + " " + str(Day)
    print(Month_Day)

    # //td/div[contains(@aria-label, "December 15")]
    Day_Xpath = '//td/div[contains(@aria-label, \"' + Month_Day + '\")]'
    print(Day_Xpath)
    ClickDay = driver.find_element(by='xpath', value=Day_Xpath)
    ClickDay.click()
    time.sleep(.5)

    DateStr = str(x18)
    Dt = DateStr
    from datetime import datetime

    # d = datetime.strptime(DateStr, "%H:%M:%S")
    # Dt = d.strftime("%I:%M %p")

    Hours = Dt[0:2]
    print(Hours)
    Min = Dt[3:5]
    print(Min)
    AMPM = Dt[6:8]
    print(AMPM)

    time.sleep(3)
    StartTime = driver.find_element(by='xpath', value='//div[@data-test="time"]//div[@data-test="startDate"]//input')
    StartTime.click()
    StartTime.send_keys(Hours)
    StartTime.send_keys(Min)
    StartTime.send_keys(AMPM)

    ScheduleType = driver.find_element(by='xpath', value='//div[@data-test="scheduleType"]//span')
    ScheduleType.click()

    scheduleTypeValue_Xpath = '//button[@title=\"' + x19 + '\"]'
    # print(ProductValue_XpathText)
    scheduleTypeValue = driver.find_element(by='xpath', value=scheduleTypeValue_Xpath)
    scheduleTypeValue.click()

    # /////////////Setting End Date///////////////////////
    EndDate = WebDriverWait(driver, 360).until(EC.presence_of_element_located((By.XPATH, '//div[@data-test="endDate"]//input[@placeholder="Choose date"]')))
    EndDate.click()
    time.sleep(0.5)
    year = x20.strftime("%Y")
    print(year)
    IRISYear = driver.find_element(by='xpath', value='//div[contains(@class, "ms-DatePicker-monthPicker monthPicker")]//div[contains(@class, "ms-DatePicker-header header")]/div').text
    print("Before Increament", IRISYear)

    while IRISYear != year:
        ClickNextYear = driver.find_element(by='xpath', value='//div[contains(@class, "ms-DatePicker-yearComponents")]//button[contains(@aria-label, "Go to next year")]')
        ClickNextYear.click()
        time.sleep(.5)
        IRISYear = driver.find_element(by='xpath', value='//div[contains(@class, "ms-DatePicker-monthPicker monthPicker")]//div[contains(@class, "ms-DatePicker-header header")]/div').text
    print("AfterIncreament", IRISYear)

    month = x20.month
    import calendar
    MonthName = calendar.month_name[month]
    print("MonthName", MonthName)

    Month_Xpath = '//div[contains(@class, "ms-DatePicker-monthPicker monthPicker")]//div[3]//button[contains(@aria-label, \"' + MonthName + '\")]'
    # print(ProductValue_XpathText)
    ClickMonth = driver.find_element(by='xpath', value=Month_Xpath)
    ClickMonth.click()
    time.sleep(.5)

    Day = x20.day
    print("Day", Day)

    Month_Day = MonthName + " " + str(Day)
    print(Month_Day)

    # //td/div[contains(@aria-label, "December 15")]
    Day_Xpath = '//td/div[contains(@aria-label, \"' + Month_Day + '\")]'
    print(Day_Xpath)
    ClickDay = driver.find_element(by='xpath', value=Day_Xpath)
    ClickDay.click()
    time.sleep(.5)

    DayOfMonth = driver.find_element(by='xpath', value='//div[@data-test="dayOfTheMonth"]')
    DayOfMonth.click()

    DayValue_Xpath = '//button[@title=\"' + str(x21) + '\"]'
    print(DayValue_Xpath)
    DayValue = driver.find_element(by='xpath', value=DayValue_Xpath)
    DayValue.click()
    time.sleep(.5)

    NextButton2.click()
    time.sleep(10)
    NextButton2.click()

    # title="1"
    time.sleep(30)
    InteractionLink = WebDriverWait(driver, 45).until(
        EC.presence_of_element_located((By.XPATH, '//div[@data-automation-key="name"]/a')))
    x24 = (InteractionLink.get_attribute('href'))

    CurrentURL = driver.current_url

    driver.quit()
    return CurrentURL, x24
#
# FilePath = "C:\\Users\\v-hcyclewala\\OneDrive - Microsoft\\V Desktop\\PythonAutomation\\IRIS_Automation\\CreateCampaign\\"
# TaskfileName = "CreateCampaign"
# Inputpath = FilePath + TaskfileName + ".xlsx"
# print(Inputpath)
#
# wb = openpyxl.load_workbook(Inputpath, data_only=True)
# ws = wb.active
#
# x1 = CampaignName = ws['A2'].value
# x2 = Line_of_Business = ws['B2'].value
# x3 = Description = ws['C2'].value
# x4 = Campaign_Owning_Product = ws['D2'].value
# x5 = Life_Cycle = ws['E2'].value
# x6 = Interactioe_Name = ws['F2'].value
# x7 = Interaction_Description = ws['G2'].value
# x8 = User_Type = ws['H2'].value
# x9 = SegmentName = ws['I2'].value
# x10 = Control = ws['J2'].value
# x11 = TreatmentName = ws['K2'].value
# x12 = Action = ws['L2'].value
# x13 = PlatformTypeValue = ws['M2'].value
# x14 = InstanceName = ws['N2'].value
# x15 = TopicId = ws['O2'].value
# x16 = Interval = ws['P2'].value
# x17 = StartDate = ws['Q2'].value
# x18 = StartTime = ws['R2'].value
# x19 = scheduleType = ws['S2'].value
# x20 = EndDate = ws['T2'].value
# x21 = DayValue = ws['U2'].value
#
# create_Campaign(CampaignName, Line_of_Business, Description, Campaign_Owning_Product, Life_Cycle, Interactioe_Name, Interaction_Description, User_Type, SegmentName, Control, TreatmentName, Action, PlatformTypeValue, InstanceName, TopicId, Interval, StartDate, StartTime, scheduleType, EndDate, DayValue)
