ADOId = "95166"
FilePath = "C:\\Users\\v-hcyclewala\\OneDrive - Microsoft\\V Desktop\\PythonAutomation\\" + ADOId + "\\IRIS_Automation\\CreateSegments\\"
TaskfileName = "IRISCreation"
Inputpath = FilePath + TaskfileName + ".xlsx"

# df = pd.read_excel(Inputpath)
# Column = df.columns.get_loc('Action')
# Column = chr(Column+65)
# print(Column)

import openpyxl

wb = openpyxl.load_workbook(Inputpath, data_only=True)
ws = wb["Campaign_Details"]
# ws = wb.active

Max_Row = ws.max_row
print(Max_Row)

InputChar = input("Press \"C\" for Create campaign \n \"I\" for Add Interaction \n \"E\" for Enabling Interaction \n \"R\" for Rescheduling")

if(InputChar == "C" ):
    from Create import CreateCamapaign as Ac

    CampaignName = ws['A2'].value
    Line_of_Business = ws['B2'].value
    Description = ws['C2'].value
    Campaign_Owning_Product = ws['D2'].value
    Life_Cycle = ws['E2'].value
    Interactioe_Name = ws['F2'].value
    Interaction_Description = ws['G2'].value
    User_Type = ws['H2'].value
    SegmentName = ws['I2'].value
    Control = ws['J2'].value
    TreatmentName = ws['K2'].value
    Action = ws['L2'].value
    PlatformTypeValue = ws['M2'].value
    InstanceName = ws['N2'].value
    TopicId = ws['O2'].value
    Interval = ws['P2'].value
    StartDate = ws['Q2'].value
    StartTime = ws['R2'].value
    scheduleType = ws['S2'].value
    EndDate = ws['T2'].value
    DayValue = ws['U2'].value

    CampaignLink, InteractionLink = Ac.create_Campaign(CampaignName, Line_of_Business, Description, Campaign_Owning_Product, Life_Cycle, Interactioe_Name, Interaction_Description, User_Type, SegmentName, Control, TreatmentName, Action, PlatformTypeValue, InstanceName, TopicId, Interval, StartDate, StartTime, scheduleType, EndDate, DayValue)
    c1 = ws.cell(row=2, column=23)
    c1.value = "Created"
    c2 = ws.cell(row=2, column=24)
    c2.value = InteractionLink
    print("CampaignCreated")
    # AddInteractionLink = CampaignLink + "/interactions/add"
    for i in range(2, Max_Row + 1):
        c1 = ws.cell(row=i, column=22)
        c1.value = CampaignLink

elif(InputChar == "I"):
    from Create import CreateInteraction as Ai
    InteractionName = []
    InteractionDescription = []
    UserType = []
    SegmentName = []
    Control = []
    TreatmentName = []
    Action = []
    PlatformType = []
    InstanceName = []
    TopicId = []
    Interval = []
    StartDate = []
    StartTime = []
    scheduleType = []
    EndDate = []
    DayOfMonth = []
    AddInteractionLink = []
    Task = []
    InteractionLink = []
    for i in range(3, Max_Row + 1):
        print(i)
        InteractionName.insert(i - 3, ws['F' + str(i)].value)
        InteractionDescription.insert(i - 3, ws['G' + str(i)].value)
        UserType.insert(i - 3, ws['H' + str(i)].value)
        SegmentName.insert(i - 3, ws['I' + str(i)].value)
        Control.insert(i - 3, ws['J' + str(i)].value)
        TreatmentName.insert(i - 3, ws['K' + str(i)].value)
        Action.insert(i - 3, ws['L' + str(i)].value)
        PlatformType.insert(i - 3, ws['M' + str(i)].value)
        InstanceName.insert(i - 3, ws['N' + str(i)].value)
        TopicId.insert(i - 3, ws['O' + str(i)].value)
        Interval.insert(i - 3, ws['P' + str(i)].value)
        StartDate.insert(i - 3, ws['Q' + str(i)].value)
        StartTime.insert(i - 3, ws['R' + str(i)].value)
        scheduleType.insert(i - 3, ws['S' + str(i)].value)
        EndDate.insert(i - 3, ws['T' + str(i)].value)
        DayOfMonth.insert(i - 3, ws['U' + str(i)].value)
        AddInteractionLink.insert(i - 3, ws['V' + str(i)].value)
        Task.insert(i - 3, ws['W' + str(i)].value)
        # InteractionLink.insert(i - 3, ws['x' + str(i)].value)
    InteractionLink, Counts = Ai.create_Interaction(InteractionName, InteractionDescription, UserType, SegmentName, Control, TreatmentName, Action,
                       PlatformType, InstanceName, TopicId, Interval, StartDate, StartTime, scheduleType, EndDate,
                       DayOfMonth, AddInteractionLink, Task)
    for i in range(len(Counts)):
        c1 = ws.cell(row=Counts[i]+3, column=24)
        c1.value = InteractionLink[i]
        c2 = ws.cell(row=Counts[i]+3, column=23)
        c2.value = "Created"
        print("Run Interaction Creation")

elif(InputChar == "E"):
    from Create import EnablingInteraction as Ei
    CampaignLink = ws['V2'].value
    InteractionName = []
    Task = []
    for i in range(2, Max_Row + 1):
        print(i)
        InteractionName.insert(i - 2, ws['F' + str(i)].value)
        Task.insert(i - 2, ws['W' + str(i)].value)
    Counts = Ei.EnableInteraction(InteractionName, CampaignLink, Task)
    for i in range(len(Counts)):
        c2 = ws.cell(row=Counts[i]+2, column=23)
        c2.value = "Enabled"
        print("Run Interaction Enable")
elif(InputChar == "R"):
    from Create import ReshedulingInteraction as Ri
    StartTime = []
    scheduleType = []
    EndDate = []
    DayOfMonth = []
    InteractionLink = []
    Task = []

    for i in range(2, Max_Row + 1):
        InteractionLink.insert(i - 2, ws['X' + str(i)].value)
        StartTime.insert(i - 2, ws['R' + str(i)].value)
        scheduleType.insert(i - 2, ws['S' + str(i)].value)
        EndDate.insert(i - 2, ws['T' + str(i)].value)
        DayOfMonth.insert(i - 2, ws['U' + str(i)].value)
        Task.insert(i - 2, ws['W' + str(i)].value)

    Ri.Reschedule_Interaction(InteractionLink, StartTime, scheduleType, EndDate, DayOfMonth, Task)
else:
    print("Enter Right Choice")

wb.save(Inputpath)