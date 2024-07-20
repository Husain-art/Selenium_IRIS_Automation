import pandas as pd
ADOId = "95166"
FilePath = "C:\\Users\\v-hcyclewala\\OneDrive - Microsoft\\V Desktop\\PythonAutomation\\" + ADOId + "\\IRIS_Automation\\CreateSegments\\"
TaskfileName = "IRISCreation"
Inputpath = FilePath + TaskfileName + ".xlsx"

ScriptPath = "C:\\Users\\v-hcyclewala\\OneDrive - Microsoft\\V Desktop\\PythonAutomation\\" + ADOId + "\\Automatic_Script_Generator\\IRISSlicer\\"

df = pd.read_excel(Inputpath)
Column = df.columns.get_loc('Action')
Column = chr(Column+65)
# print(Column)

import CreateSegment as cs
import EditSegment as es
import openpyxl

wb = openpyxl.load_workbook(Inputpath)
ws = wb.active

Max_Row = ws.max_row
Task = []
Segment_Name = []
Products = []
Descriptions = []
Delay_In_Publish = []
Audi_Tag = []
Camp_Tag = []
Freq_Tag = []
User_Type = []
Segment_Link = []
FileName = []

for i in range(2, Max_Row+1):
    Action_Index = Column+str(i)
    Task.insert(i-2, ws['A' + str(i)].value)
    Segment_Name.insert(i-2, ws['B' + str(i)].value)
    Products.insert(i - 2, ws['C' + str(i)].value)
    Descriptions.insert(i - 2, ws['D' + str(i)].value)
    Delay_In_Publish.insert(i - 2, ws['E' + str(i)].value)
    Audi_Tag.insert(i - 2, ws['F' + str(i)].value)
    Camp_Tag.insert(i - 2, ws['G' + str(i)].value)
    Freq_Tag.insert(i - 2, ws['H' + str(i)].value)
    User_Type.insert(i - 2, ws['I' + str(i)].value)
    Segment_Link.insert(i - 2, ws['J' + str(i)].value)
    FileName.insert(i - 2, ws['K' + str(i)].value)

print(Task)
if "Create" in Task:
    print(Task)
    SegmentLinks, Counts = cs.create_Segment(Task, Segment_Name, Products, Descriptions, Delay_In_Publish, Audi_Tag, Camp_Tag, Freq_Tag, User_Type, FileName, ScriptPath)
    for i in range(len(Counts)):
        c1 = ws.cell(row=Counts[i]+2, column=10)
        c1.value = SegmentLinks[i]
        c2 = ws.cell(row=Counts[i]+2, column=1)
        c2.value = "Created"
        print("Run Segment Creation")

if "Edit" in Task:
    Counts = es.edit_Segment(Task, Segment_Name, Descriptions, Delay_In_Publish, Audi_Tag, Camp_Tag, Freq_Tag, Segment_Link, FileName, ScriptPath)
    for i in range(len(Counts)):
        c2 = ws.cell(row=Counts[i]+2, column=1)
        c2.value = "Edited"
        print("Edit Segment Run")

wb.save(Inputpath)
